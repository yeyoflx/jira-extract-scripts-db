import jira.client
from jira.client import JIRA
import openpyxl
from openpyxl import Workbook
import datetime
from datetime import date
import time
import json

now = datetime.datetime.now()



def extract_from_jira(jira,project):
    #Used to search for specfici JIRA issues based on issue type
    issues1 = jira.search_issues('project = '+project+' AND issuetype = Story ORDER BY created DESC',startAt = 0  ,maxResults=100)
    issues2 = jira.search_issues('project = '+project+' AND issuetype = Story ORDER BY created DESC',startAt = 100,maxResults=500)
    issues3 = jira.search_issues('project = '+project+' AND issuetype = Story ORDER BY created DESC',startAt = 200,maxResults=500)

    list_of_searches = [issues1,
                        issues2,
                        issues3
                        ]
    dict_of_stories = {}
    # Accesses JIRA items changelog and locates changes in status as well as save the date of the changed status
    for list_of_issues in list_of_searches:
        for issue2 in list_of_issues:
            print(issue2)
            issue = jira.issue(str(issue2), expand = 'changelog')
            changelog = issue.changelog
            last_status = issue.fields.created[0:10].split("-")
            first_status = date(int(last_status[0]), int(last_status[1]), int(last_status[2]))
            date_created = date(int(last_status[0]), int(last_status[1]), int(last_status[2]))
            total_days = 0

            #used to determine Number of days from date user sotry was created to todays run date
            todays_date = date(now.year,now.month,now.day)
            datetime.timedelta(7)
            story_age = (todays_date-date_created).days

            #used to determine Date the user story 3was assigned the Current Workflow
            #Also used to determine Number of dats from date user story was assigned to the current to todays run date
            second_date = 0
            for history in changelog.histories[::-1]:
                for item in history.items:
                    if item.field == "status":
                        history_created = history.created[0:10].split("-")
                        second_date = date(int(history_created[0]), int(history_created[1]), int(history_created[2]))
                        datetime.timedelta(7)
                        total_days += (second_date - first_status).days
                        first = second_date
            if second_date == 0:
                second_date = date_created
                status_age = (todays_date - date_created).days
            else:
                status_age = (todays_date - second_date).days

            # Puts values from JIRA api and changelog into a dictionary containing all JIRA issues
            if str(issue2) not in dict_of_stories.keys():
                dict_of_stories[str(issue)] = {"Summary":issue2.fields.summary,
                                               "Priority":issue2.fields.priority.name,
                                               "Reporter":issue2.fields.reporter.displayName,
                                               "Date Created":str(date_created),
                                               "Story Age": str(story_age),
                                               "Assignee":issue2.fields.assignee.displayName,
                                               "Current Status":issue2.fields.status.name,
                                               "Status Date": str(second_date),
                                               "Status Age": status_age
                                               }
    return dict_of_stories





def check_statuses(jira,project):

    final_dict ={}
    #Used to search for specfici JIRA issues based on issue type
    issues1 = jira.search_issues('project = ' + project + ' AND issuetype = Story ORDER BY created DESC', startAt=0,   maxResults=100)
    issues2 = jira.search_issues('project = ' + project + ' AND issuetype = Story ORDER BY created DESC', startAt=100, maxResults=500)
    issues3 = jira.search_issues('project = ' + project + ' AND issuetype = Story ORDER BY created DESC', startAt=200, maxResults=500)

    list_of_searches = [issues1,
                        issues2,
                        issues3
                        ]


    list_of_statuses = ["Recording","Technical Assessment","Prioritization","Planning","Definition","Design","Ready for Development", "Development","Testing on DEV","Ready for Production","Closed"]
    # Stores the status and dates for each issue and all the statuses it has passed through
    for list_of_issues in list_of_searches:
        for issue2 in list_of_issues:
            if str(issue2) not in final_dict:
                dict_of_statuses = {}
                cleaned_dict = {}
                print(issue2)
                issue = jira.issue(str(issue2), expand='changelog')
                changelog = issue.changelog
                for history in changelog.histories[::-1]:
                    for item in history.items:
                        if item.field == "status":
                            #print("Date:" + history.created + " From: " + item.fromString + " To:" + item.toString)
                            if item.fromString == "Recording":
                                dict_of_statuses[item.fromString] = [("First date:",issue2.fields.created)]
                            if item.fromString in dict_of_statuses:
                                dict_of_statuses[item.fromString].append(("Last date:",history.created))
                            if item.fromString not in dict_of_statuses:
                                dict_of_statuses[item.fromString] = [("First date:",history.created)]
                            if item.toString not in dict_of_statuses:
                                dict_of_statuses[item.toString] = [("First date:",history.created)]
                # Compares the issues statuses to all possible statuses to create a dictionary containing all statuses for one issue
                for status in list_of_statuses:
                    if status in dict_of_statuses:
                        cleaned_dict[status] = dict_of_statuses[status]
                    else:
                        cleaned_dict[status] = None
                for k,v in cleaned_dict.items():
                    if v != None:
                        a = len(v)
                        if len(v) == 1:
                            first = v[0][1][0:10].split("-")
                            date1 = date(int(first[0]), int(first[1]), int(first[2]))
                            #print(k,date1)
                            cleaned_dict[k] = [date1,0]
                        else:
                            first = v[0][1][0:10].split("-")
                            second = v[1][1][0:10].split("-")
                            datetime.timedelta(7)
                            date1 = date(int(first[0]),int(first[1]),int(first[2]))
                            date2 = date(int(second[0]),int(second[1]),int(second[2]))
                            difference = (date2-date1).days
                            #print(k,date1 ,date2,difference)
                            cleaned_dict[k] = [date1, date2, difference]
                    else:
                        pass
                        #print("None")
                #print(cleaned_dict)
                if str(issue2) not in final_dict:
                    final_dict[str(issue2)] = cleaned_dict
    return final_dict

"""
second_date = None
for history in changelog.histories[::-1]:
    for item in history.items:
        if item.field =="status":
            history_created = history.created[0:10].split("-")
            second_date = date(int(history_created[0]), int(history_created[1]), int(history_created[2]))
            datetime.timedelta(7)
            print( "Date:" + history.created + " From: " + item.fromString + " To:" + item.toString + " Total days between: " + str((second_date-date_created).days))
            total_days += (second_date - date_created).days
            date_created = second_date
"""





def write_to_workbook(dict_of_stories,worksheet):
    # Writes the stored issue and its attributes as well as its current status onto an Excel workbook
    header_names = ["Issue","Summary","Priority","Reporter","Date Created","Assignee","Current Status","Status Date","Status Age"]
    headers = ["A","B","C","D","E","F","G","H","I"]


    column_name = 0
    row_count = 1
    for column in headers:
        worksheet[column+str(row_count)] = header_names[column_name]
        column_name += 1

    column_name = 0
    row_count = 2
    for k,v in dict_of_stories.items():
        for column2 in headers:
            if headers[column_name] == "A":
                worksheet[column2+str(row_count)] = k
                column_name += 1
            else:
                cell = column2+str(row_count)
                print(cell, column_name)
                worksheet[cell] = v[header_names[column_name]]
                column_name += 1
        column_name = 0
        row_count +=1

def write_to_workbook2(dict_of_statuses,worksheet):
    # Writes the issues and all statuses into an excel file
    headers = ["A", "B", "C", "D", "E", "F", "G", "H", "I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AG","AH","AI"]
    column = 0
    row = 2
    for k,v in dict_of_statuses.items():
        worksheet[headers[column]+str(row)] = k
        column += 1
        print(v)
        for k1, v1 in v.items():
            print(k1,v1)
            if v1 != None:
                if len(v1) == 3:
                    for i in v1:
                        print(headers[column] + str(row))
                        worksheet[headers[column] + str(row)] = i
                        column += 1
                else:
                    v1.insert(1,"None")
                    for i in v1:
                        print(headers[column] + str(row))
                        worksheet[headers[column] + str(row)] = i
                        column += 1
            else:
                for i in range(3):
                    print(i)
                    print(headers[column] + str(row))
                    worksheet[headers[column] + str(row)] = "None"
                    column += 1
        column = 0
        row += 1




if __name__ == '__main__':
    start_time = time.time()
    # Used to connect to the Schenker JIRA Database using personal credentials
    options = {'server': 'https://schenkereservices.atlassian.net'}
    jira = JIRA(options, basic_auth=('Diego.Felix@DBSchenker.com', 'V0lkswagen00151637?'))
    print("Successfully connected to JIRA")

    # Creates new excel workbook
    wb = Workbook()
    dest_filename = 'C:\\Users\yeyoflx\Documents\JIRA Data Visualization Project\JIRA PM Reporting\\jira_extraction2.xlsx'
    projects = ["EFO", "EFA", "EFL"]
    ws1 = wb.create_sheet(projects[0])
    ws2 = wb.create_sheet(projects[1])
    ws3 = wb.create_sheet(projects[2])

    ws1.title = projects[0]
    ws2.title = projects[1]
    ws3.title = projects[2]
    work_sheets = [ws1,ws2,ws3]
    for i in range(len(work_sheets)):
        dict_of_statuses = check_statuses(jira,projects[i])
        write_to_workbook2(dict_of_statuses,work_sheets[i])
        wb.save(filename=dest_filename)
    """
    for i in range(3):
        dict_of_stories = extract_from_jira(jira,projects[i])
        write_to_workbook(dict_of_stories,work_sheets[i])
        wb.save(filename=dest_filename)
    """
    print(time.time() - start_time, 'seconds it took to run')